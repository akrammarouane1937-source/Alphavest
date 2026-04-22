"""
Daily Automation Script — Alphavest PFE
========================================
Run every morning after the price CSV is updated with yesterday's closing prices.

What it does:
  1. Reads the last available date in the price file -- that is END_DATE (yesterday's close)
  2. Computes rolling windows: END_DATE - 1Y/2Y/3Y/4Y/5Y
  3. Recomputes all 10 risk metrics for all 77 stocks + MASI
  4. Recomputes all 10 portfolio metrics (MASI-weighted)
  5. Overwrites both Excel output files

Schedule via Windows Task Scheduler to run at 08:00 every morning.
"""

import pandas as pd
import numpy as np
import warnings
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

# ============================================================
# CONFIGURATION
# ============================================================
BASE_DIR          = r'c:\Users\àf\Downloads\School project'
PRICES_FILE       = os.path.join(BASE_DIR, 'historical_prices.csv')
RF_FILE           = os.path.join(BASE_DIR, 'taux_sans_risque_maroc_quotidien.csv')
WEIGHTS_FILE      = os.path.join(BASE_DIR, 'Compo_All_Indices_20260408.xls')
OUTPUT_STOCKS     = os.path.join(BASE_DIR, 'risk_metrics_by_window_latest.xlsx')
OUTPUT_PORTFOLIO  = os.path.join(BASE_DIR, 'portfolio_metrics_latest.xlsx')

TRADING_DAYS = 252
MIN_OBS      = 30

# ============================================================
# STEP 1 — LOAD DATA
# ============================================================
print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Loading data...")

prices = pd.read_csv(PRICES_FILE, sep=';', dayfirst=True,
                     parse_dates=['Date'], index_col='Date', dtype=object)
prices.index = pd.to_datetime(prices.index, dayfirst=True, errors='coerce')
for col in prices.columns:
    prices[col] = prices[col].astype(str).str.replace(',', '.', regex=False).str.strip()
prices = prices.apply(pd.to_numeric, errors='coerce').sort_index()

rf_raw = pd.read_csv(RF_FILE, sep=';', parse_dates=['Date'],
                     index_col='Date', encoding='utf-8-sig')
rf_annual_decimal = rf_raw['BDT_52_semaines_%'] / 100.0
rf_daily = ((1 + rf_annual_decimal) ** (1.0 / TRADING_DAYS) - 1).sort_index()

masi_prices  = prices['MASI']
stock_prices = prices.drop(columns=['MASI'])
stocks       = list(stock_prices.columns)

# ============================================================
# STEP 2 — DYNAMIC ROLLING WINDOWS
# END_DATE = last date in the price file (yesterday's close after morning update)
# ============================================================
END_DATE = prices.index.max()

WINDOWS = {
    '1Y': (END_DATE - pd.DateOffset(years=1), END_DATE),
    '2Y': (END_DATE - pd.DateOffset(years=2), END_DATE),
    '3Y': (END_DATE - pd.DateOffset(years=3), END_DATE),
    '4Y': (END_DATE - pd.DateOffset(years=4), END_DATE),
    '5Y': (END_DATE - pd.DateOffset(years=5), END_DATE),
}

print(f"  END_DATE  = {END_DATE.date()}  (last available closing price)")
for label, (s, e) in WINDOWS.items():
    print(f"  {label}: {s.date()} to {e.date()}")

# ============================================================
# STEP 3 — DAILY RETURNS
# ============================================================
print("Computing daily returns...")
returns_dict = {}
for col in stock_prices.columns:
    valid = stock_prices[col].dropna()
    returns_dict[col] = valid.pct_change()
all_returns = pd.DataFrame(returns_dict)

masi_valid   = masi_prices.dropna()
masi_returns = masi_valid.pct_change()

# ============================================================
# METRIC FUNCTIONS (identical formulas to risk_metrics_engine_v2.py)
# ============================================================
def slice_win(series, start, end):
    return series.loc[start:end].dropna()

def calc_annualized_return(r):
    n = len(r)
    if n < MIN_OBS:
        return np.nan
    return float((1 + r).prod() ** (TRADING_DAYS / n) - 1)

def calc_annualized_vol(r):
    if len(r) < MIN_OBS:
        return np.nan
    return float(r.std() * np.sqrt(TRADING_DAYS))

def calc_annualized_variance(r):
    if len(r) < MIN_OBS:
        return np.nan
    return float(r.var() * TRADING_DAYS)

def calc_rf_annual(r_index):
    rf_aligned = rf_daily.reindex(r_index, method='ffill')
    return float((1 + rf_aligned).prod() - 1)

def calc_sharpe(r_ann, rf_ann, vol):
    if any(np.isnan(x) for x in [r_ann, rf_ann, vol]) or vol == 0:
        return np.nan
    return float((r_ann - rf_ann) / vol)

def calc_beta(r, r_m):
    common = r.index.intersection(r_m.index)
    if len(common) < MIN_OBS:
        return np.nan
    rc, rmc = r.loc[common], r_m.loc[common]
    var_m = rmc.var()
    if var_m == 0:
        return np.nan
    return float(rc.cov(rmc) / var_m)

def calc_tracking_error(r, r_m):
    common = r.index.intersection(r_m.index)
    if len(common) < MIN_OBS:
        return np.nan
    return float((r.loc[common] - r_m.loc[common]).std() * np.sqrt(TRADING_DAYS))

def calc_alpha(r_ann, rf_ann, beta, r_masi_ann):
    if any(np.isnan(x) for x in [r_ann, rf_ann, beta, r_masi_ann]):
        return np.nan
    return float(r_ann - (rf_ann + beta * (r_masi_ann - rf_ann)))

def calc_max_drawdown(r):
    if len(r) < MIN_OBS:
        return np.nan
    cum = (1 + r).cumprod()
    return float(abs(((cum - cum.cummax()) / cum.cummax()).min()))

def calc_var95(r):
    if len(r) < MIN_OBS:
        return np.nan
    return float(-np.percentile(r.dropna(), 5))

def calc_cvar95(r):
    if len(r) < MIN_OBS:
        return np.nan
    var = calc_var95(r)
    tail = r[r <= -var]
    return float(-tail.mean()) if len(tail) > 0 else np.nan

def calc_sortino(r, r_ann, rf_ann):
    if len(r) < MIN_OBS or np.isnan(r_ann):
        return np.nan
    dd = np.sqrt(np.mean(np.minimum(r, 0) ** 2)) * np.sqrt(TRADING_DAYS)
    return float((r_ann - rf_ann) / dd) if dd > 0 else np.nan

def compute_metrics(r, r_masi_win):
    R_ann  = calc_annualized_return(r)
    vol    = calc_annualized_vol(r)
    rf_ann = calc_rf_annual(r.index)
    R_m    = calc_annualized_return(r_masi_win)
    return {
        'Annualized_Variance':   calc_annualized_variance(r),
        'Annualized_Volatility': vol,
        'Annualized_Return':     R_ann,
        'Sharpe_Ratio':          calc_sharpe(R_ann, rf_ann, vol),
        'Beta':                  calc_beta(r, r_masi_win),
        'Tracking_Error':        calc_tracking_error(r, r_masi_win),
        'Alpha':                 calc_alpha(R_ann, rf_ann, calc_beta(r, r_masi_win), R_m),
        'Max_Drawdown':          calc_max_drawdown(r),
        'VaR_95':                calc_var95(r),
        'CVaR_95':               calc_cvar95(r),
        'Sortino_Ratio':         calc_sortino(r, R_ann, rf_ann),
    }

_test_idx = pd.date_range('2024-01-01', periods=40, freq='B')
METRIC_NAMES = list(compute_metrics(pd.Series([0.01, -0.01] * 20, index=_test_idx),
                                    pd.Series([0.01, -0.01] * 20, index=_test_idx)).keys())
PCT_METRICS  = {'Annualized_Variance', 'Annualized_Volatility', 'Annualized_Return',
                'Tracking_Error', 'Max_Drawdown', 'VaR_95', 'CVaR_95'}

# ============================================================
# STEP 4 — COMPUTE INDIVIDUAL STOCK + MASI METRICS
# ============================================================
print("Computing individual stock metrics...")
all_tickers = ['MASI'] + stocks
stock_results = {m: pd.DataFrame(index=all_tickers,
                                  columns=list(WINDOWS.keys()), dtype=float)
                 for m in METRIC_NAMES}

for wlabel, (wstart, wend) in WINDOWS.items():
    r_masi_win = slice_win(masi_returns, wstart, wend)
    R_masi     = calc_annualized_return(r_masi_win)

    # MASI row
    m = compute_metrics(r_masi_win, r_masi_win)
    for metric, val in m.items():
        stock_results[metric].loc['MASI', wlabel] = val

    # Each stock
    for stock in stocks:
        r = slice_win(all_returns[stock], wstart, wend)
        m = compute_metrics(r, r_masi_win)
        for metric, val in m.items():
            stock_results[metric].loc[stock, wlabel] = val

    print(f"  [{wlabel}] done — {len(r_masi_win)} MASI trading days")

# ============================================================
# STEP 5 — PORTFOLIO WEIGHTS (equal weight across all 80 stocks)
# ============================================================
# Equal weight: each stock gets 1/N where N = number of stocks in price history
_all_stocks = list(all_returns.columns)
_eq_weight  = 1.0 / len(_all_stocks)
weights = {t: _eq_weight for t in _all_stocks}

if False:  # kept for reference only — old MASI weights logic
 NAME_TO_TICKER = {
    'ATTIJARIWAFA BANK':             'ATW.CS',
    'SODEP-Marsa Maroc':             'MSA.CS',
    'MANAGEM':                       'MNG.CS',
    'ITISSALAT AL-MAGHRIB':          'IAM.CS',
    'LafargeHolcim Maroc':           'LHM.CS',
    'BANK OF AFRICA':                'BOA.CS',
    'TGCC S.A':                      'TGC.CS',
    'CIMENTS DU MAROC':              'CMT.CS',
    'SGTM S.A':                      None,
    'AKDITAL':                       'AKT.CS',
    'BCP':                           'BCP.CS',
    'TAQA MOROCCO':                  'TQM.CS',
    'CFG BANK':                      'CFG.CS',
    'COSUMAR':                       'CSR.CS',
    'LABEL VIE':                     'LBV.CS',
    'DOUJA PROM ADDOHA':             'ADH.CS',
    'WAFA ASSURANCE':                'WAA.CS',
    'AFRIQUIA GAZ':                  'GAZ.CS',
    'ALLIANCES':                     'ADI.CS',
    'MINIERE TOUISSIT':              None,
    'CIH':                           'CIH.CS',
    'CMGP GROUP':                    'CMG.CS',
    'ARADEI CAPITAL':                'ARD.CS',
    'LESIEUR CRISTAL':               'LES.CS',
    'CDM':                           'CDM.CS',
    'JET CONTRACTORS':               'JET.CS',
    'SONASID':                       'SID.CS',
    'HPS':                           'HPS.CS',
    'TOTALENERGIES MARKETING MAROC': 'TMA.CS',
    'SOTHEMA':                       'SOT.CS',
    'SMI':                           'SMI.CS',
    'RESIDENCES DAR SAADA':          'RDS.CS',
    'MUTANDIS SCA':                  'MUT.CS',
    'VICENNE':                       None,
    'ATLANTASANAD':                  'ATL.CS',
    'BMCI':                          'BCI.CS',
    'DELTA HOLDING':                 'DHO.CS',
    'CASH PLUS S.A':                 'CRS.CS',
    'SOCIETE DES BOISSONS DU MAROC': 'SBM.CS',
    'SANLAM MAROC':                  'SAM.CS',
    'RISMA':                         'RIS.CS',
    'AUTO HALL':                     'ATH.CS',
    'MICRODATA':                     'MIC.CS',
    'IMMORENTE INVEST':              'IMO.CS',
    'DISWAY':                        'DWY.CS',
    'AFMA':                          'AFM.CS',
    'AGMA':                          'AGM.CS',
    'EQDOM':                         'EQD.CS',
    'COLORADO':                      'COL.CS',
    'SALAFIN':                       'SLF.CS',
    'SNEP':                          'SNP.CS',
    'DARI COUSPATE':                 'DRI.CS',
    'OULMES':                        'OUL.CS',
    'UNIMER':                        'UMR.CS',
    'ALUMINIUM DU MAROC':            'ALM.CS',
    'DISTY TECHNOLOGIES':            'DYT.CS',
    'AUTO NEJMA':                    'NEJ.CS',
    'ENNAKL':                        'NKL.CS',
    'CTM':                           'CTM.CS',
    'MAGHREBAIL':                    'MAB.CS',
    'FENIE BROSSETTE':               'FBR.CS',
    'PROMOPHARM S.A.':               'PRO.CS',
    'S.M MONETIQUE':                 'S2M.CS',
    'STOKVIS NORD AFRIQUE':          'SNA.CS',
    'MAROC LEASING':                 'MLE.CS',
    'M2M Group':                     'M2M.CS',
    'BALIMA':                        'BAL.CS',
    'STROC INDUSTRIE':               'STR.CS',
    'AFRIC INDUSTRIES SA':           'AFI.CS',
    'MAGHREB OXYGENE':               'MOX.CS',
    'CARTIER SAADA':                 'CMA.CS',
    'REALISATIONS MECANIQUES':       'SRM.CS',
    'MED PAPER':                     'MDP.CS',
    'INVOLYS':                       'INV.CS',
    'DELATTRE LEVIVIER MAROC':       'DLM.CS',
    'ZELLIDJA S.A':                  'ZDJ.CS',
    'IB MAROC.COM':                  'IBC.CS',
    'REBAB COMPANY':                 'REB.CS',
 }

# ============================================================
# STEP 6 — PORTFOLIO DAILY RETURNS
# ============================================================
print("Computing portfolio daily returns...")
valid_tickers = [t for t in weights if t in all_returns.columns]
w_series = pd.Series({t: weights[t] for t in valid_tickers})

portfolio_returns = pd.Series(index=all_returns.index, dtype=float)
for date in all_returns.index:
    row = all_returns.loc[date, valid_tickers]
    available = row.dropna()
    if len(available) == 0:
        portfolio_returns[date] = np.nan
        continue
    w_norm = w_series[available.index] / w_series[available.index].sum()
    portfolio_returns[date] = (w_norm * available).sum()
portfolio_returns = portfolio_returns.dropna()

# ============================================================
# STEP 7 — PORTFOLIO METRICS
# ============================================================
print("Computing portfolio metrics...")
port_results = {}
masi_bench   = {}

for wlabel, (wstart, wend) in WINDOWS.items():
    r_p   = slice_win(portfolio_returns, wstart, wend)
    r_m   = slice_win(masi_returns, wstart, wend)
    port_results[wlabel] = compute_metrics(r_p, r_m)
    masi_bench[wlabel]   = compute_metrics(r_m, r_m)

# ============================================================
# EXCEL STYLING HELPERS
# ============================================================
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
ALT_FILL    = PatternFill("solid", fgColor="EEF2F7")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
PORT_FILL   = PatternFill("solid", fgColor="1B5E20")
MASI_FILL   = PatternFill("solid", fgColor="4A1942")

HEADER_FONT = Font(name='Calibri', bold=True, color="FFFFFF", size=11)
PORT_FONT   = Font(name='Calibri', bold=True, color="FFFFFF", size=11)
MASI_FONT   = Font(name='Calibri', bold=True, color="FFFFFF", size=11)
DATA_FONT   = Font(name='Calibri', size=10)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left',   vertical='center')

thin_side   = Side(style='thin', color="CCCCCC")
THIN_BORDER = Border(left=thin_side, right=thin_side,
                     top=thin_side,  bottom=thin_side)

def style(cell, fill, font, align, border):
    cell.fill = fill; cell.font = font
    cell.alignment = align; cell.border = border

def header_cell(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=val)
    style(c, HEADER_FILL, HEADER_FONT, CENTER, THIN_BORDER)

def data_cell(ws, row, col, val, row_fill):
    c = ws.cell(row=row, column=col, value=val)
    style(c, row_fill, DATA_FONT, CENTER, THIN_BORDER)

# ============================================================
# STEP 8 — WRITE STOCK METRICS EXCEL
# ============================================================
print(f"Writing {OUTPUT_STOCKS}...")
wb1 = Workbook()

win_labels  = list(WINDOWS.keys())
win_display = {l: f"{l}  ({WINDOWS[l][0].date()} to {WINDOWS[l][1].date()})"
               for l in win_labels}

# Sheet: Daily_Returns
ws_dr = wb1.active
ws_dr.title = 'Daily_Returns'
dr_headers = ['Date'] + stocks
for ci, val in enumerate(dr_headers, 1):
    header_cell(ws_dr, 1, ci, val)

# Build daily returns DataFrame for display (5Y window only, to keep it manageable)
r_start_5y = WINDOWS['5Y'][0]
display_dates = sorted([d for d in all_returns.index if d >= r_start_5y])
for ri, date in enumerate(display_dates, 2):
    rf = ALT_FILL if ri % 2 == 0 else WHITE_FILL
    data_cell(ws_dr, ri, 1, date.strftime('%d/%m/%Y'), rf)
    for ci, stock in enumerate(stocks, 2):
        val = all_returns.loc[date, stock]
        data_cell(ws_dr, ri, ci, round(float(val)*100, 6) if not np.isnan(val) else None, rf)

ws_dr.column_dimensions['A'].width = 14
ws_dr.freeze_panes = 'B2'

# One sheet per metric
METRIC_DISPLAY = {
    'Annualized_Variance':   'Annualized Variance (%)',
    'Annualized_Volatility': 'Annualized Volatility (%)',
    'Annualized_Return':     'Annualized Return (%)',
    'Sharpe_Ratio':          'Sharpe Ratio',
    'Beta':                  'Beta',
    'Tracking_Error':        'Tracking Error (%)',
    'Alpha':                 "Jensen's Alpha",
    'Max_Drawdown':          'Max Drawdown (%)',
    'VaR_95':                'VaR 95% (%)',
    'CVaR_95':               'CVaR 95% (%)',
    'Sortino_Ratio':         'Sortino Ratio',
}

for metric in METRIC_NAMES:
    is_pct = metric in PCT_METRICS
    ws = wb1.create_sheet(title=metric[:31])

    headers = ['Ticker'] + [win_display[l] for l in win_labels]
    for ci, val in enumerate(headers, 1):
        header_cell(ws, 1, ci, val)

    for ri, ticker in enumerate(all_tickers, 2):
        rf = MASI_FILL if ticker == 'MASI' else (ALT_FILL if ri % 2 == 0 else WHITE_FILL)
        fnt = MASI_FONT if ticker == 'MASI' else DATA_FONT
        c = ws.cell(row=ri, column=1, value=ticker)
        style(c, rf, fnt, CENTER, THIN_BORDER)
        for ci, wl in enumerate(win_labels, 2):
            v = stock_results[metric].loc[ticker, wl]
            val = round(float(v)*100, 4) if is_pct and not np.isnan(float(v)) else (round(float(v), 4) if not np.isnan(float(v)) else None)
            c = ws.cell(row=ri, column=ci, value=val)
            style(c, rf, fnt, CENTER, THIN_BORDER)

    ws.column_dimensions['A'].width = 14
    for ci in range(2, len(win_labels)+2):
        ws.column_dimensions[get_column_letter(ci)].width = 30
    ws.freeze_panes = 'B2'

wb1.save(OUTPUT_STOCKS)

# ============================================================
# STEP 9 — WRITE PORTFOLIO METRICS EXCEL
# ============================================================
print(f"Writing {OUTPUT_PORTFOLIO}...")
wb2 = Workbook()

# Sheet 1: Portfolio_Returns (full history, single column)
ws_ret = wb2.active
ws_ret.title = 'Portfolio_Returns'
header_cell(ws_ret, 1, 1, 'Date')
header_cell(ws_ret, 1, 2, 'Portfolio Return (%)')
for ri, (date, ret) in enumerate(portfolio_returns.sort_index().items(), 2):
    rf = ALT_FILL if ri % 2 == 0 else WHITE_FILL
    data_cell(ws_ret, ri, 1, date.strftime('%d/%m/%Y'), rf)
    data_cell(ws_ret, ri, 2, round(ret*100, 6), rf)
ws_ret.column_dimensions['A'].width = 16
ws_ret.column_dimensions['B'].width = 22
ws_ret.freeze_panes = 'A2'

# Sheet 2: All_Metrics (all 10 metrics, Portfolio vs MASI, one sheet)
ws_m = wb2.create_sheet(title='All_Metrics')
col_headers = ['Metric', 'Entity'] + [win_display[l] for l in win_labels]
for ci, val in enumerate(col_headers, 1):
    header_cell(ws_m, 1, ci, val)

ws_m.column_dimensions['A'].width = 26
ws_m.column_dimensions['B'].width = 28
for ci in range(3, len(win_labels)+3):
    ws_m.column_dimensions[get_column_letter(ci)].width = 30
ws_m.freeze_panes = 'C2'

row_idx = 2
for metric in METRIC_NAMES:
    is_pct = metric in PCT_METRICS
    name   = METRIC_DISPLAY[metric]

    # MASI row
    c = ws_m.cell(row=row_idx, column=1, value=name)
    style(c, MASI_FILL, MASI_FONT, LEFT, THIN_BORDER)
    c = ws_m.cell(row=row_idx, column=2, value='MASI (Benchmark)')
    style(c, MASI_FILL, MASI_FONT, CENTER, THIN_BORDER)
    for ci, wl in enumerate(win_labels, 3):
        v = masi_bench[wl][metric]
        val = round(v*100, 4) if is_pct and not np.isnan(v) else (round(v, 4) if not np.isnan(v) else None)
        c = ws_m.cell(row=row_idx, column=ci, value=val)
        style(c, MASI_FILL, MASI_FONT, CENTER, THIN_BORDER)
    row_idx += 1

    # Portfolio row
    c = ws_m.cell(row=row_idx, column=1, value=name)
    style(c, PORT_FILL, PORT_FONT, LEFT, THIN_BORDER)
    c = ws_m.cell(row=row_idx, column=2, value='Portfolio (MASI-weighted)')
    style(c, PORT_FILL, PORT_FONT, CENTER, THIN_BORDER)
    for ci, wl in enumerate(win_labels, 3):
        v = port_results[wl][metric]
        val = round(v*100, 4) if is_pct and not np.isnan(v) else (round(v, 4) if not np.isnan(v) else None)
        c = ws_m.cell(row=row_idx, column=ci, value=val)
        style(c, PORT_FILL, PORT_FONT, CENTER, THIN_BORDER)
    row_idx += 2  # blank separator between metric groups

wb2.save(OUTPUT_PORTFOLIO)

# ============================================================
# DONE
# ============================================================
run_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
print(f"\n[{run_time}] Done.")
print(f"  END_DATE used    : {END_DATE.date()}")
print(f"  Stocks computed  : {len(stocks)}")
print(f"  Portfolio stocks : {len(valid_tickers)}")
print(f"  Output (stocks)  : {OUTPUT_STOCKS}")
print(f"  Output (portfolio): {OUTPUT_PORTFOLIO}")
