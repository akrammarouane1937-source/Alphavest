"""
Portfolio Metrics Engine — Alphavest PFE
Computes all 10 risk metrics for the MASI-weighted portfolio
across 5 rolling windows (1Y/2Y/3Y/4Y/5Y) ending 17/03/2026.

Weights source: Compo_All_Indices_20260408.xls (MASI sheet)
Stocks not in price history (VICENNE, ENNAKL, REALISATIONS MECANIQUES)
have their weights redistributed proportionally to the 75 mapped stocks.
"""

import pandas as pd
import numpy as np
import warnings
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

# ============================================================
# CONFIGURATION
# ============================================================
BASE_DIR     = r'c:\Users\àf\Downloads\School project'
PRICES_FILE  = os.path.join(BASE_DIR, 'historical_prices.csv')
RF_FILE      = os.path.join(BASE_DIR, 'taux_sans_risque_maroc_quotidien.csv')
WEIGHTS_FILE = os.path.join(BASE_DIR, 'Compo_All_Indices_20260408.xls')
OUTPUT_FILE  = os.path.join(BASE_DIR, 'portfolio_metrics_by_window_v3.xlsx')

TRADING_DAYS = 252
MIN_OBS      = 30

WINDOWS = {
    '1Y': ('2025-03-17', '2026-03-17'),
    '2Y': ('2024-03-18', '2026-03-17'),
    '3Y': ('2023-03-17', '2026-03-17'),
    '4Y': ('2022-03-17', '2026-03-17'),
    '5Y': ('2021-03-17', '2026-03-17'),
}

METRIC_NAMES = [
    'Annualized_Variance',
    'Annualized_Volatility',
    'Annualized_Return',
    'Sharpe_Ratio',
    'Beta',
    'Tracking_Error',
    'Alpha',
    'Max_Drawdown',
    'VaR_95',
    'CVaR_95',
    'Sortino_Ratio',
]

PCT_METRICS = {
    'Annualized_Variance', 'Annualized_Volatility', 'Annualized_Return',
    'Tracking_Error', 'Max_Drawdown', 'VaR_95', 'CVaR_95'
}

# ============================================================
# NAME -> TICKER MAPPING  (MASI company name -> .CS ticker)
# ============================================================
NAME_TO_TICKER = {
    'ATTIJARIWAFA BANK':             'ATW.CS',
    'SODEP-Marsa Maroc':             'MSA.CS',
    'MANAGEM':                       'MNG.CS',
    'ITISSALAT AL-MAGHRIB':          'IAM.CS',
    'LafargeHolcim Maroc':           'LHM.CS',
    'BANK OF AFRICA':                'BOA.CS',
    'TGCC S.A':                      'TGC.CS',
    'CIMENTS DU MAROC':              'CMT.CS',
    'SGTM S.A':                      None,    # not in price history
    'AKDITAL':                       'AKT.CS',
    'BCP':                           'BCP.CS',
    'TAQA MOROCCO':                  'TQM.CS',
    'CFG BANK':                      'CFG.CS',
    'COSUMAR':                       'CSR.CS',
    'LABEL VIE':                     'LBV.CS',
    'DOUJA PROM ADDOHA':             'ADH.CS',
    'WAFA ASSURANCE':                'WAA.CS',
    'AFRIQUIA GAZ':                  'GAZ.CS',
    'ALLIANCES':                     'ADI.CS',
    'MINIERE TOUISSIT':              None,    # not in price history

    'CIH':                           'CIH.CS',
    'CMGP GROUP':                    'CMG.CS',
    'ARADEI CAPITAL':                'ARD.CS',
    'LESIEUR CRISTAL':               'LES.CS',
    'CDM':                           'CDM.CS',
    'JET CONTRACTORS':               'JET.CS',
    'SONASID':                       'SID.CS',
    'HPS':                           'HPS.CS',
    'TOTALENERGIES MARKETING MAROC': 'TMA.CS',
    'SOTHEMA':                       'SOT.CS',
    'SMI':                           'SMI.CS',
    'RESIDENCES DAR SAADA':          'RDS.CS',
    'MUTANDIS SCA':                  'MUT.CS',
    'VICENNE':                       None,   # not in price history
    'ATLANTASANAD':                  'ATL.CS',
    'BMCI':                          'BCI.CS',
    'DELTA HOLDING':                 'DHO.CS',
    'CASH PLUS S.A':                 'CRS.CS',
    'SOCIETE DES BOISSONS DU MAROC': 'SBM.CS',
    'SANLAM MAROC':                  'SAM.CS',
    'RISMA':                         'RIS.CS',
    'AUTO HALL':                     'ATH.CS',
    'MICRODATA':                     'MIC.CS',
    'IMMORENTE INVEST':              'IMO.CS',
    'DISWAY':                        'DWY.CS',
    'AFMA':                          'AFM.CS',
    'AGMA':                          'AGM.CS',
    'EQDOM':                         'EQD.CS',
    'COLORADO':                      'COL.CS',
    'SALAFIN':                       'SLF.CS',
    'SNEP':                          'SNP.CS',
    'DARI COUSPATE':                 'DRI.CS',
    'OULMES':                        'OUL.CS',
    'UNIMER':                        'UMR.CS',
    'ALUMINIUM DU MAROC':            'ALM.CS',
    'DISTY TECHNOLOGIES':            'DYT.CS',
    'AUTO NEJMA':                    'NEJ.CS',
    'ENNAKL':                        'NKL.CS',
    'CTM':                           'CTM.CS',
    'MAGHREBAIL':                    'MAB.CS',
    'FENIE BROSSETTE':               'FBR.CS',
    'PROMOPHARM S.A.':               'PRO.CS',
    'S.M MONETIQUE':                 'S2M.CS',
    'STOKVIS NORD AFRIQUE':          'SNA.CS',
    'MAROC LEASING':                 'MLE.CS',
    'M2M Group':                     'M2M.CS',
    'BALIMA':                        'BAL.CS',
    'STROC INDUSTRIE':               'STR.CS',
    'AFRIC INDUSTRIES SA':           'AFI.CS',
    'MAGHREB OXYGENE':               'MOX.CS',
    'CARTIER SAADA':                 'CMA.CS',
    'REALISATIONS MECANIQUES':       'SRM.CS',
    'MED PAPER':                     'MDP.CS',
    'INVOLYS':                       'INV.CS',
    'DELATTRE LEVIVIER MAROC':       'DLM.CS',
    'ZELLIDJA S.A':                  'ZDJ.CS',
    'IB MAROC.COM':                  'IBC.CS',
    'REBAB COMPANY':                 'REB.CS',
}

# ============================================================
# LOAD & BUILD WEIGHTS
# ============================================================
print("Loading weights from MASI composition file...")
masi_compo = pd.read_excel(WEIGHTS_FILE, sheet_name='MASI')
masi_compo['Instrument_clean'] = masi_compo['Instrument'].str.strip()
masi_compo['Ticker'] = masi_compo['Instrument_clean'].map(NAME_TO_TICKER)

# Stocks not in price history → their weight will be redistributed
unmapped = masi_compo[masi_compo['Ticker'].isna()]
mapped   = masi_compo[masi_compo['Ticker'].notna()].copy()

print(f"  Mapped: {len(mapped)} stocks  |  Unmapped (no price history): {len(unmapped)} stocks")
print(f"  Unmapped stocks: {unmapped['Instrument_clean'].tolist()}")
print(f"  Unmapped total weight: {unmapped['Poids'].sum():.4%}")

# Redistribute unmapped weights proportionally to mapped stocks
total_mapped_weight = mapped['Poids'].sum()
mapped['Weight_adjusted'] = mapped['Poids'] / total_mapped_weight  # renormalise to 1.0

print(f"  Adjusted weights sum: {mapped['Weight_adjusted'].sum():.6f}")

# Build weights dict: ticker -> adjusted weight
weights = dict(zip(mapped['Ticker'], mapped['Weight_adjusted']))

# ============================================================
# LOAD PRICE & RF DATA
# ============================================================
print("\nLoading price data...")
prices = pd.read_csv(PRICES_FILE, sep=';', dayfirst=True,
                     parse_dates=['Date'], index_col='Date', dtype=object)
for col in prices.columns:
    prices[col] = prices[col].astype(str).str.replace(',', '.', regex=False).str.strip()
prices = prices.apply(pd.to_numeric, errors='coerce').sort_index()

print("Loading risk-free rate data...")
rf_raw = pd.read_csv(RF_FILE, sep=';', parse_dates=['Date'],
                     index_col='Date', encoding='utf-8-sig')
rf_annual_decimal = rf_raw['BDT_52_semaines_%'] / 100.0
rf_daily = (1 + rf_annual_decimal) ** (1.0 / TRADING_DAYS) - 1
rf_daily = rf_daily.sort_index()

masi_prices  = prices['MASI']
stock_prices = prices.drop(columns=['MASI'])

print(f"  Price range: {prices.index[0].date()} -> {prices.index[-1].date()}")

# ============================================================
# COMPUTE DAILY RETURNS (per stock, dropna first)
# ============================================================
print("Computing daily returns...")
returns_dict = {}
for col in stock_prices.columns:
    valid = stock_prices[col].dropna()
    returns_dict[col] = valid.pct_change()
all_returns = pd.DataFrame(returns_dict)

masi_valid   = masi_prices.dropna()
masi_returns = masi_valid.pct_change()

# ============================================================
# COMPUTE PORTFOLIO DAILY RETURNS
# r_p,t = sum(w_i * r_i,t)   for all stocks that have a return on day t
# On days where some stocks have no return (NaN), we use the sub-portfolio
# of stocks that did trade, scaled so weights still sum to 1.
# ============================================================
print("Computing portfolio daily returns...")

# All dates where at least one stock traded
all_dates = all_returns.index

portfolio_returns = pd.Series(index=all_dates, dtype=float)

# Filter weights to only stocks present in our returns df
valid_tickers = [t for t in weights if t in all_returns.columns]
w_series = pd.Series({t: weights[t] for t in valid_tickers})

print(f"  Portfolio stocks with price data: {len(valid_tickers)} / {len(weights)}")

for date in all_dates:
    row = all_returns.loc[date, valid_tickers]
    available = row.dropna()
    if len(available) == 0:
        portfolio_returns[date] = np.nan
        continue
    # Renormalise weights for available stocks on this date
    w_avail = w_series[available.index]
    w_norm  = w_avail / w_avail.sum()
    portfolio_returns[date] = (w_norm * available).sum()

portfolio_returns = portfolio_returns.dropna()
print(f"  Portfolio return dates: {len(portfolio_returns)} ({portfolio_returns.index[0].date()} -> {portfolio_returns.index[-1].date()})")

# ============================================================
# HELPER FUNCTIONS (same formulas as risk_metrics_engine_v2.py)
# ============================================================
def slice_win(series, start, end):
    return series.loc[start:end].dropna()

def calc_annualized_return(r):
    n = len(r)
    if n < MIN_OBS:
        return np.nan
    return float((1 + r).prod() ** (TRADING_DAYS / n) - 1)

def calc_annualized_vol(r):
    if len(r) < MIN_OBS:
        return np.nan
    return float(r.std() * np.sqrt(TRADING_DAYS))

def calc_annualized_variance(r):
    if len(r) < MIN_OBS:
        return np.nan
    return float(r.var() * TRADING_DAYS)

def calc_rf_annual(r_index):
    rf_aligned = rf_daily.reindex(r_index, method='ffill')
    return float((1 + rf_aligned).prod() - 1)

def calc_sharpe(r_ann, rf_ann, vol):
    if any(np.isnan(x) for x in [r_ann, rf_ann, vol]) or vol == 0:
        return np.nan
    return float((r_ann - rf_ann) / vol)

def calc_beta(r, r_m):
    common = r.index.intersection(r_m.index)
    if len(common) < MIN_OBS:
        return np.nan
    rc, rmc = r.loc[common], r_m.loc[common]
    var_m = rmc.var()
    if var_m == 0:
        return np.nan
    return float(rc.cov(rmc) / var_m)

def calc_tracking_error(r, r_m):
    common = r.index.intersection(r_m.index)
    if len(common) < MIN_OBS:
        return np.nan
    active = r.loc[common] - r_m.loc[common]
    return float(active.std() * np.sqrt(TRADING_DAYS))

def calc_alpha(r_ann, rf_ann, beta, r_masi_ann):
    if any(np.isnan(x) for x in [r_ann, rf_ann, beta, r_masi_ann]):
        return np.nan
    return float(r_ann - (rf_ann + beta * (r_masi_ann - rf_ann)))

def calc_max_drawdown(r):
    if len(r) < MIN_OBS:
        return np.nan
    cum = (1 + r).cumprod()
    peak = cum.cummax()
    dd = (cum - peak) / peak
    return float(abs(dd.min()))

def calc_var95(r):
    if len(r) < MIN_OBS:
        return np.nan
    return float(-np.percentile(r, 5))

def calc_cvar95(r):
    if len(r) < MIN_OBS:
        return np.nan
    var = calc_var95(r)
    tail = r[r <= -var]
    if len(tail) == 0:
        return np.nan
    return float(-tail.mean())

def calc_sortino(r, r_ann, rf_ann):
    if len(r) < MIN_OBS or np.isnan(r_ann) or np.isnan(rf_ann):
        return np.nan
    downside = r[r < 0]
    if len(downside) == 0:
        return np.nan
    downside_std = np.sqrt((downside ** 2).mean()) * np.sqrt(TRADING_DAYS)
    if downside_std == 0:
        return np.nan
    return float((r_ann - rf_ann) / downside_std)

def compute_all_metrics(r, r_masi_win, label='Portfolio'):
    R_ann    = calc_annualized_return(r)
    variance = calc_annualized_variance(r)
    vol      = calc_annualized_vol(r)
    rf_ann   = calc_rf_annual(r.index)
    sharpe   = calc_sharpe(R_ann, rf_ann, vol)
    beta     = calc_beta(r, r_masi_win)
    te       = calc_tracking_error(r, r_masi_win)
    R_m_ann  = calc_annualized_return(r_masi_win)
    alpha    = calc_alpha(R_ann, rf_ann, beta, R_m_ann)
    mdd      = calc_max_drawdown(r)
    var      = calc_var95(r)
    cvar     = calc_cvar95(r)
    sortino  = calc_sortino(r, R_ann, rf_ann)

    return {
        'Annualized_Variance':   variance,
        'Annualized_Volatility': vol,
        'Annualized_Return':     R_ann,
        'Sharpe_Ratio':          sharpe,
        'Beta':                  beta,
        'Tracking_Error':        te,
        'Alpha':                 alpha,
        'Max_Drawdown':          mdd,
        'VaR_95':                var,
        'CVaR_95':               cvar,
        'Sortino_Ratio':         sortino,
    }

# ============================================================
# COMPUTE METRICS ACROSS ALL WINDOWS
# ============================================================
print("\nComputing portfolio metrics across all windows...")
results = {}  # metric -> {window: value}

for win_label, (start, end) in WINDOWS.items():
    r_p   = slice_win(portfolio_returns, start, end)
    r_m   = slice_win(masi_returns, start, end)

    print(f"  {win_label}: portfolio={len(r_p)} obs, masi={len(r_m)} obs")

    metrics = compute_all_metrics(r_p, r_m, label=f'Portfolio {win_label}')
    for metric, value in metrics.items():
        if metric not in results:
            results[metric] = {}
        results[metric][win_label] = value

# Also compute MASI benchmark metrics for comparison
print("\nComputing MASI benchmark metrics across all windows...")
masi_results = {}
for win_label, (start, end) in WINDOWS.items():
    r_m = slice_win(masi_returns, start, end)
    metrics = compute_all_metrics(r_m, r_m, label=f'MASI {win_label}')
    for metric, value in metrics.items():
        if metric not in masi_results:
            masi_results[metric] = {}
        masi_results[metric][win_label] = value

# Print summary table
print("\n" + "="*70)
print("PORTFOLIO vs MASI — SUMMARY")
print("="*70)
win_labels = list(WINDOWS.keys())
header = f"{'Metric':<25}" + "".join(f"{'Port_'+w:>12}" for w in win_labels) + "".join(f"{'MASI_'+w:>12}" for w in win_labels)
print(header)
print("-"*70)
for metric in METRIC_NAMES:
    row = f"{metric:<25}"
    for w in win_labels:
        v = results[metric].get(w, np.nan)
        if metric in PCT_METRICS:
            row += f"{v*100:>11.2f}%" if not np.isnan(v) else f"{'N/A':>12}"
        else:
            row += f"{v:>12.4f}" if not np.isnan(v) else f"{'N/A':>12}"
    for w in win_labels:
        v = masi_results[metric].get(w, np.nan)
        if metric in PCT_METRICS:
            row += f"{v*100:>11.2f}%" if not np.isnan(v) else f"{'N/A':>12}"
        else:
            row += f"{v:>12.4f}" if not np.isnan(v) else f"{'N/A':>12}"
    print(row)

# ============================================================
# EXCEL OUTPUT
# ============================================================
print("\nWriting Excel output...")

HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
ALT_FILL     = PatternFill("solid", fgColor="EEF2F7")
PORT_FILL    = PatternFill("solid", fgColor="1B5E20")  # dark green for portfolio
MASI_FILL    = PatternFill("solid", fgColor="4A1942")  # dark purple for MASI
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT  = Font(name='Calibri', bold=True, color="FFFFFF", size=11)
PORT_FONT    = Font(name='Calibri', bold=True, color="FFFFFF", size=11)
MASI_FONT    = Font(name='Calibri', bold=True, color="FFFFFF", size=11)
DATA_FONT    = Font(name='Calibri', size=10)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN   = Alignment(horizontal='left', vertical='center')

thin = Side(style='thin', color="CCCCCC")
THIN_BORDER  = Border(left=thin, right=thin, top=thin, bottom=thin)

METRIC_DISPLAY = {
    'Annualized_Variance':   'Annualized Variance',
    'Annualized_Volatility': 'Annualized Volatility',
    'Annualized_Return':     'Annualized Return',
    'Sharpe_Ratio':          'Sharpe Ratio',
    'Beta':                  'Beta',
    'Tracking_Error':        'Tracking Error',
    'Alpha':                 "Jensen's Alpha",
    'Max_Drawdown':          'Max Drawdown',
    'VaR_95':                'VaR 95%',
    'CVaR_95':               'CVaR 95%',
    'Sortino_Ratio':         'Sortino Ratio',
}

wb = Workbook()
win_labels = list(WINDOWS.keys())

# -----------------------------------------------
# Sheet 1: Portfolio_Returns  (single column, full history)
# -----------------------------------------------
ws_ret = wb.active
ws_ret.title = 'Portfolio_Returns'

ret_headers = ['Date', 'Portfolio Return (%)']
for col_idx, val in enumerate(ret_headers, start=1):
    cell = ws_ret.cell(row=1, column=col_idx, value=val)
    cell.fill = HEADER_FILL; cell.font = HEADER_FONT
    cell.alignment = CENTER_ALIGN; cell.border = THIN_BORDER

# Full portfolio return history (all dates, not sliced by window)
full_returns = portfolio_returns.sort_index()
for row_idx, (date, ret) in enumerate(full_returns.items(), start=2):
    fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL
    ws_ret.cell(row=row_idx, column=1, value=date.strftime('%d/%m/%Y')).fill = fill
    ws_ret.cell(row=row_idx, column=1).font = DATA_FONT
    ws_ret.cell(row=row_idx, column=1).alignment = CENTER_ALIGN
    ws_ret.cell(row=row_idx, column=1).border = THIN_BORDER
    ws_ret.cell(row=row_idx, column=2, value=round(ret * 100, 6)).fill = fill
    ws_ret.cell(row=row_idx, column=2).font = DATA_FONT
    ws_ret.cell(row=row_idx, column=2).alignment = CENTER_ALIGN
    ws_ret.cell(row=row_idx, column=2).border = THIN_BORDER

ws_ret.column_dimensions['A'].width = 16
ws_ret.column_dimensions['B'].width = 22
ws_ret.freeze_panes = 'A2'

# -----------------------------------------------
# Sheet 2: All_Metrics  (all 10 metrics, Portfolio vs MASI)
# -----------------------------------------------
ws = wb.create_sheet(title='All_Metrics')

# Column headers: Metric | Entity | 1Y | 2Y | 3Y | 4Y | 5Y
col_headers = ['Metric', 'Entity'] + win_labels
for col_idx, val in enumerate(col_headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=val)
    cell.fill = HEADER_FILL; cell.font = HEADER_FONT
    cell.alignment = CENTER_ALIGN; cell.border = THIN_BORDER

ws.column_dimensions['A'].width = 24
ws.column_dimensions['B'].width = 26
for col_idx in range(3, len(win_labels) + 3):
    ws.column_dimensions[get_column_letter(col_idx)].width = 14

row_idx = 2
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
SECTION_FONT = Font(name='Calibri', bold=True, size=10, color="1F3864")

for metric in METRIC_NAMES:
    is_pct = metric in PCT_METRICS
    unit_suffix = ' (%)' if is_pct else ''
    display_name = METRIC_DISPLAY[metric] + unit_suffix

    # --- MASI row ---
    ws.cell(row=row_idx, column=1, value=display_name)
    ws.cell(row=row_idx, column=1).fill = MASI_FILL
    ws.cell(row=row_idx, column=1).font = MASI_FONT
    ws.cell(row=row_idx, column=1).alignment = LEFT_ALIGN
    ws.cell(row=row_idx, column=1).border = THIN_BORDER

    ws.cell(row=row_idx, column=2, value='MASI (Benchmark)')
    ws.cell(row=row_idx, column=2).fill = MASI_FILL
    ws.cell(row=row_idx, column=2).font = MASI_FONT
    ws.cell(row=row_idx, column=2).alignment = CENTER_ALIGN
    ws.cell(row=row_idx, column=2).border = THIN_BORDER

    for col_idx, win in enumerate(win_labels, start=3):
        v = masi_results[metric].get(win, np.nan)
        val = round(v * 100, 4) if is_pct and not np.isnan(v) else (round(v, 4) if not np.isnan(v) else None)
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = MASI_FILL; cell.font = MASI_FONT
        cell.alignment = CENTER_ALIGN; cell.border = THIN_BORDER
    row_idx += 1

    # --- Portfolio row ---
    ws.cell(row=row_idx, column=1, value=display_name)
    ws.cell(row=row_idx, column=1).fill = PORT_FILL
    ws.cell(row=row_idx, column=1).font = PORT_FONT
    ws.cell(row=row_idx, column=1).alignment = LEFT_ALIGN
    ws.cell(row=row_idx, column=1).border = THIN_BORDER

    ws.cell(row=row_idx, column=2, value='Portfolio (MASI-weighted)')
    ws.cell(row=row_idx, column=2).fill = PORT_FILL
    ws.cell(row=row_idx, column=2).font = PORT_FONT
    ws.cell(row=row_idx, column=2).alignment = CENTER_ALIGN
    ws.cell(row=row_idx, column=2).border = THIN_BORDER

    for col_idx, win in enumerate(win_labels, start=3):
        v = results[metric].get(win, np.nan)
        val = round(v * 100, 4) if is_pct and not np.isnan(v) else (round(v, 4) if not np.isnan(v) else None)
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = PORT_FILL; cell.font = PORT_FONT
        cell.alignment = CENTER_ALIGN; cell.border = THIN_BORDER
    row_idx += 1

    # blank separator row between metric groups
    row_idx += 1

ws.freeze_panes = 'C2'

wb.save(OUTPUT_FILE)
print(f"\nSaved: {OUTPUT_FILE}")

# -----------------------------------------------
# Print weights used
# -----------------------------------------------
print("\nTop 15 portfolio weights (adjusted):")
sorted_weights = sorted(weights.items(), key=lambda x: -x[1])
for ticker, w in sorted_weights[:15]:
    name = masi_compo.set_index('Ticker').loc[ticker, 'Instrument_clean'] if ticker in masi_compo['Ticker'].values else ticker
    print(f"  {ticker:10s} {w:.4%}  ({name})")
print(f"  ... and {len(weights)-15} more")
print(f"\nTotal weight check: {sum(weights.values()):.8f}")
